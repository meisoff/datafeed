import json
import datetime
import openpyxl as ox
from openpyxl.styles import Font
from collections import Counter
import crud
import os

own_base = None
url_base = None

# with open("own_base.json", "r", encoding="UTF-8") as f:
#     own_base = json.load(f)
# with open("url_base.json", "r", encoding="UTF-8") as f:
#     url_base = json.load(f)


# ГЕНЕРИРУЮ ФЕЙКОВЫЕ РЕЗУЛЬТАТЫ ДЛЯ ПРОВЕРКИ
# НА ДАННОМ ЭТАПЕ ДОЛЖЕН ПОЛУЧАТЬ ФАЙЛ СЛЕДУЮЩЕГО ТИПА:
#         "index": int
#         "date": date
#         "url": str
#         "price": int,
#         "stock": bool
#     }

result_base = list(crud.get_data_for_file())
product_base = list(crud.get_from_products())

# for item in url_base:
#     result_base.append({
#         "index": item["index"],
#         "date": datetime.datetime.now(),
#         "url": item["url"],
#         "price": round(eval("100" + item["formula"]), 2) if item["formula"] else 100,  # Пробуем использование формулы
#         "stock": True if int(item["index"]) % 10 else False  # Генерируем наличие
#     })

wb = ox.Workbook()
ws = wb.active
ws.title = "result"

# Записываем основные стобцы
own_title = ["Код", "Бренд", "Название", "МВЦ руб.", "MIN руб.", "AVG руб.", "MAX руб."]

for i in range(len(own_title)):
    ws.cell(1, i + 1).value = own_title[i]

# Считаем максимальное количество ссылок на товар
all_indexes = []
for item in result_base:
    all_indexes.append(item["index"])
max_count = max([values for item, values in Counter(all_indexes).items()])

# Добавляем стобцы для ссылок
for i in range(max_count):
    ws.cell(1, 8 + i).value = f"item_{i + 1}"

for i in range(len(product_base)):
    count = 8  # counter of column

    ws.cell(i + 2, 1).value = product_base[i]["index"]
    ws.cell(i + 2, 2).value = product_base[i]["brand"]
    ws.cell(i + 2, 3).value = product_base[i]["name"]
    ws.cell(i + 2, 4).value = product_base[i]["mpp"]

    min = 9999999999
    max = 0
    avg = []

    for j in range(len(result_base)):
        if product_base[i]["index"] == result_base[j]["index"]:
            price = int(result_base[j]["price"]) if result_base[j]["price"] else 0
            ws.cell(i + 2, count).hyperlink = result_base[j]["link"]
            ws.cell(i + 2, count).value = price
            if not result_base[j]["stock"]:
                ws.cell(i + 2, count).font = Font(color="616161")

            if price > max:
                max = price
            if price < min:
                min = price
            avg.append(price)

            count += 1

    ws.cell(i + 2, 5).value = min if len(avg) else None
    ws.cell(i + 2, 6).value = round(sum(avg) / len(avg), 2) if len(avg) else None
    ws.cell(i + 2, 7).value = max if len(avg) else None

date = datetime.datetime.now().strftime('%Y-%m-%d__%H-%M-%S')
name = f"{date}_result.xlsx"
wb.save(name)
size = os.path.getsize(name)
#TODO: Сделать добавление данных в fileinfo
crud.create_file_info({
    "date": date,
    "name": name,
    "size": size,
    "format": "xlsx"
})